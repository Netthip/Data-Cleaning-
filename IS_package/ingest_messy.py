# -*- coding: utf-8 -*-
"""
ingest_messy.py
อ่านไฟล์ Excel รายงานดิบ (มีบรรยาย/หัวหลายบรรทัด) แล้วดึงเฉพาะ "ตาราง" ให้แบน
ผลลัพธ์รวมไว้ที่ data/clean_data.xlsx
"""
import os, re
import pandas as pd
from openpyxl import load_workbook
import yaml

BASE = os.getcwd()
DATA = os.path.join(BASE, "data")
CFG  = os.path.join(BASE, "config", "ingest_headers.yml")
OUT  = os.path.join(BASE, "data", "clean_data.xlsx")

def normalize(s):
    s = "" if s is None else str(s)
    return re.sub(r"\s+", " ", s.replace("\n"," ").strip())

def load_cfg():
    with open(CFG,"r",encoding="utf-8") as f:
        return yaml.safe_load(f)

def load_datadict(cfg):
    dd = (cfg or {}).get("datadict", {})
    path = dd.get("path")
    if not path or not os.path.exists(path):
        return {}
    df = pd.read_excel(path, sheet_name=dd.get("sheet",0)).fillna("")
    mapping = {}
    for _,r in df.iterrows():
        std = normalize(r.iloc[0])
        if not std: 
            continue
        aliases = [normalize(x) for x in r.iloc[1:].tolist() if normalize(x)]
        mapping.setdefault(std, set()).update(aliases)
    return {k:list(v) for k,v in mapping.items()}

def map_cols(cols, dd_map, rename_final):
    out = []
    for c in cols:
        cn = normalize(c)
        mapped = None
        # datadict exact / alias
        for std, aliases in dd_map.items():
            if cn == std or cn in aliases:
                mapped = std; break
        if not mapped:
            low = cn.lower()
            # heuristic ปี/บัญชี
            if "2569" in low or "fy2569" in low or "fy69" in low or re.search(r"[^0-9]69[^0-9]", low):
                mapped = "ปี2569"
            elif "2570" in low or "fy2570" in low or "fy70" in low or re.search(r"[^0-9]70[^0-9]", low):
                mapped = "ปี2570"
            elif re.search(r"(บัญชี|account|หมวด)\s*1", low): mapped = "บัญชี1"
            elif re.search(r"(บัญชี|account|หมวด)\s*2", low): mapped = "บัญชี2"
            elif re.search(r"(บัญชี|account|หมวด)\s*3", low): mapped = "บัญชี3"
            elif any(k in low for k in ["หน่วยงาน","หน่วยรับงบ","สำนัก","กอง"]): mapped = "หน่วยงาน"
            elif any(k in low for k in ["กิจกรรม","activity","งาน"]): mapped = "กิจกรรม"
            elif any(k in low for k in ["ประจำ","ลงทุน","ประเภทค่าใช้จ่าย","ลักษณะค่าใช้จ่าย"]): mapped = "ประจำ/ลงทุน"
            elif any(k in low for k in ["ผูกพัน","ปีเดียว"]): mapped = "ปีเดียว/ผูกพัน"
            elif any(k in low for k in ["ผลผลิต","โครงการ"]): mapped = "ผลผลิต/โครงการ"
        out.append(mapped or cn)
    # rename_final
    out = [rename_final.get(c,c) for c in out]
    # ทำให้ไม่ซ้ำ
    seen = {}
    uniq = []
    for n in out:
        if n not in seen: seen[n]=0; uniq.append(n)
        else: seen[n]+=1; uniq.append(f"{n}_{seen[n]}")
    return uniq

def detect_header_row(rows, hints, max_scan):
    best_i, best_score = 0, -1
    for i in range(min(len(rows), max_scan)):
        row = [normalize(x) for x in rows[i]]
        nonempty = sum(1 for x in row if x not in ("","nan"))
        score = nonempty
        if any(h in " ".join(row) for h in hints):
            score += 5
        if score > best_score:
            best_score, best_i = score, i
    return best_i

def read_table(path, cfg, dd_map):
    EXTRA_INGEST_IGNORES = ['datadict_FULL.xlsx', 'datadic.xlsx', 'national_budget_datadict.xlsx', 'แบบฟอร์ม_2569.xlsx', 'แบบฟอร์ม_2569_filled.xlsx', 'clean_data.xlsx', 'clean_with_formulas.xlsx', 'clean_with_formulas_agg.xlsx']
    ignore = set((cfg or {}).get("ignore_files", [])) | set(EXTRA_INGEST_IGNORES)
    base = os.path.basename(path)
    if base in ignore:
        print("[INFO] ข้ามไฟล์ตาม ignore:", base)
        return pd.DataFrame()

    ovs = (cfg or {}).get("file_overrides", {})
    ov  = ovs.get(base, {})
    sheet = ov.get("sheet")  # อาจ None
    header_row = ov.get("header_row")
    header_span= int(ov.get("header_span", 1))

    wb = load_workbook(path, data_only=True, read_only=True)
    sh = wb[sheet] if sheet in wb.sheetnames else wb[wb.sheetnames[0]]

    # ดึงค่าทั้งชีตเป็น list of rows
    rows = [list(r) for r in sh.iter_rows(values_only=True)]
    if not rows:
        return pd.DataFrame()

    if header_row is None:
        hints = (cfg.get("header_hints") or {}).get("keywords", ["รหัส","รายการ","งบ","ปี","หน่วยงาน","กิจกรรม"])
        max_scan = (cfg.get("header_hints") or {}).get("max_scan", 120)
        header_row = detect_header_row(rows, hints, max_scan)

    # ผสานหัวหลายบรรทัด (เช่น 2 บรรทัด)
    head = [normalize(x) for x in rows[header_row]]
    if header_span >= 2 and header_row+1 < len(rows):
        nxt = [normalize(x) for x in rows[header_row+1]]
        L = max(len(head), len(nxt))
        head = [(head[i] if i<len(head) and head[i] else "") + (" " + nxt[i] if i<len(nxt) and nxt[i] else "") for i in range(L)]
        head = [normalize(x) for x in head]

    data_rows = rows[header_row+header_span:]
    # จำกัดให้เท่าความยาวหัว
    width = len(head)
    fixed = []
    for r in data_rows:
        rr = list(r) + [None]*max(0, width-len(r))
        fixed.append(rr[:width])
    df = pd.DataFrame(fixed, columns=head)

    # ตัดคอลัมน์ว่างล้วน
    df = df.loc[:, df.notna().any(axis=0)]

    # map ชื่อคอลัมน์
    rename_final = (cfg or {}).get("rename_final", {})
    df.columns = map_cols(df.columns.tolist(), dd_map, rename_final)

    # ลบแถวว่างล้วน
    df = df.dropna(how="all")
    # แปลงตัวเลขพื้นฐาน
    numeric_like = set((cfg or {}).get("numeric_like", []))
    percent_like = set((cfg or {}).get("percent_like", []))
    for c in df.columns:
        s = df[c]
        if s.dtype==object:
            ss = s.astype(str).str.replace(",","",regex=False)
            # ถ้าเป็น % ที่อยู่ใน percent_like ให้ตัดเครื่องหมาย %
            if c in percent_like:
                ss = ss.str.replace("%","",regex=False)
            num = pd.to_numeric(ss, errors="coerce")
            if num.notna().sum() >= max(5, int(len(df)*0.02)):
                df[c] = num.astype("Float64")

    df["__source_file"] = base
    return df

def main():
    cfg = load_cfg()
    dd_map = load_datadict(cfg)
    all_frames = []
    for fn in os.listdir(DATA):
        if not fn.lower().endswith((".xlsx",".xls")): 
            continue
        path = os.path.join(DATA, fn)
        try:
            part = read_table(path, cfg, dd_map)
            if not part.empty:
                all_frames.append(part)
                print(f"[OK] {fn} -> rows:{len(part)} cols:{len(part.columns)}")
            else:
                print(f"[WARN] {fn} -> ว่าง/ข้าม")
        except Exception as e:
            print(f"[ERR ] {fn} ->", e)

    if not all_frames:
        print("[ERROR] ไม่มีข้อมูลหลัง ingest (ตรวจ ignore, header_row, sheet)")
        return

    # รวม + เขียน clean_data.xlsx
    cols_union = []
    for f in all_frames:
        for c in f.columns:
            if c not in cols_union:
                cols_union.append(c)
    ALL = pd.concat([f.reindex(columns=cols_union) for f in all_frames], ignore_index=True)

    # เตือนถ้าคอลัมน์หลักขาด
    expected = set((cfg or {}).get("expected_core", []))
    if expected:
        missing = [c for c in expected if c not in ALL.columns]
        if missing:
            print("[WARN] คอลัมน์หลักขาด:", missing)

    ALL.to_excel(OUT, index=False)
    print("[OK] wrote clean table ->", OUT, "rows:", len(ALL), "cols:", len(ALL.columns))

if __name__ == "__main__":
    main()
