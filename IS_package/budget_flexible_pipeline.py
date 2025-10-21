# -*- coding: utf-8 -*-
"""
budget_flexible_pipeline_FIXED.py
เวอร์ชันนี้แก้ indent ทั้งหมดให้เรียบร้อย
และมีคำอธิบายภาษาไทยในแต่ละส่วน
"""

import os
import re
import pandas as pd
import numpy as np
import yaml
from openpyxl import load_workbook
try:
    # สำหรับ openpyxl เวอร์ชันใหม่
    from openpyxl.utils.cell import coordinate_from_string, column_index_from_string, get_column_letter
except ImportError:
    # เผื่อใช้เวอร์ชันเก่า
    from openpyxl.utils import coordinate_from_string, column_index_from_string, get_column_letter


# -------------------------------------------------------------------
# พื้นฐาน path สำหรับอ่าน/เขียนไฟล์
# -------------------------------------------------------------------
BASE_DIR = os.getcwd()
DATA_DIR = os.path.join(BASE_DIR, "data")
OUT_DIR = os.path.join(BASE_DIR, "out")
CFG_PATH = os.path.join(BASE_DIR, "config", "formulas_budget_workflow.yml")

os.makedirs(DATA_DIR, exist_ok=True)
os.makedirs(OUT_DIR, exist_ok=True)

# -------------------------------------------------------------------
# กำหนดชื่อคอลัมน์มาตรฐานกับคำพ้อง
# -------------------------------------------------------------------
COLUMN_SYNONYMS = {
    "หน่วยงาน": ["หน่วยงาน", "กรม", "สถาบัน"],
    "แผนงาน": ["แผนงาน", "แผน"],
    "ผลผลิต/โครงการ": ["ผลผลิต", "โครงการ", "โครงการ/ผลผลิต"],
    "กิจกรรม": ["กิจกรรม", "งาน"],
    "รายการ": ["รายการ", "รายละเอียดรายการ"],
    "ประจำ/ลงทุน": ["ประจำ/ลงทุน", "ลักษณะค่าใช้จ่าย", "ประเภทค่าใช้จ่าย"],
    "ปีเดียว/ผูกพัน": ["ปีเดียว/ผูกพัน", "ลักษณะผูกพัน"],
    "ปี2569": ["ปี2569", "ปี 2569", "FY2569"],
    "ปี2570": ["ปี2570", "ปี 2570", "FY2570"],
    "บัญชี1": ["บัญชี1", "บัญชี 1"],
    "บัญชี2": ["บัญชี2", "บัญชี 2"],
    "บัญชี3": ["บัญชี3", "บัญชี 3"]
}

PREFERRED_SHEETS = ["5รายละเอียด", "รายละเอียด5", "รายละเอียด"]

# -------------------------------------------------------------------
# ฟังก์ชันย่อย
# -------------------------------------------------------------------
def normalize_text(s):
    if s is None:
        return ""
    s = str(s).replace("\n", " ").strip()
    return re.sub(r"\s+", " ", s)


def detect_header_row(df, max_scan=25):
    """สแกนหาบรรทัดหัวตารางโดยดูจากคำที่คุ้นเคย"""
    hints = ["รหัส", "รายการ", "งบ", "ปี", "หน่วยงาน"]
    m = min(len(df), max_scan)
    for i in range(m):
        row = df.iloc[i].astype(str).tolist()
        if any(h in " ".join(row) for h in hints):
            return i
    return 0


def make_unique(names):
    """ทำให้ชื่อคอลัมน์ไม่ซ้ำ"""
    seen = {}
    out = []
    for n in names:
        if n not in seen:
            seen[n] = 0
            out.append(n)
        else:
            seen[n] += 1
            out.append(f"{n}_{seen[n]}")
    return out


def map_columns_flex(df, synonyms):
    """รีแมปชื่อคอลัมน์ให้ตรงกับ datadict"""
    new_cols = []
    for c in df.columns:
        c_norm = normalize_text(c)
        mapped = None
        for std, cands in synonyms.items():
            if c_norm == std or c_norm in cands:
                mapped = std
                break
        new_cols.append(mapped if mapped else c_norm)
    df.columns = make_unique(new_cols)
    return df


def eval_expr_df(df, expr):
    """ใช้ eval คำนวณสูตรในคอลัมน์"""
    ns = {str(c): pd.to_numeric(df[c], errors="ignore") for c in df.columns}
    ns.update({
        "sum": lambda x: pd.to_numeric(x, errors="coerce").sum(skipna=True),
        "mean": lambda x: pd.to_numeric(x, errors="coerce").mean(skipna=True),
    })
    return eval(expr, {"__builtins__": {}}, ns)


def safe_divide(a, b):
    """หารแบบปลอดภัย"""
    a = pd.to_numeric(a, errors="coerce")
    b = pd.to_numeric(b, errors="coerce")
    out = a / b
    return out.replace([np.inf, -np.inf], np.nan)


# -------------------------------------------------------------------
# อ่าน datadict จาก Excel
# -------------------------------------------------------------------
def load_datadict_to_synonyms(cfg, synonyms):
    dd = cfg.get("datadict", {})
    path = dd.get("path")
    if not path or not os.path.exists(path):
        print("[WARN] datadict path ไม่พบ:", path)
        return synonyms

    try:
        df = pd.read_excel(path, sheet_name=dd.get("sheet", 0)).fillna("")
    except Exception as e:
        print("[WARN] อ่าน datadict ไม่สำเร็จ:", e)
        return synonyms

    newmap = {}
    if df.shape[1] >= 2:
        for _, r in df.iterrows():
            std = normalize_text(r.iloc[0])
            aliases = [normalize_text(x) for x in r.iloc[1:].tolist() if x]
            newmap.setdefault(std, []).extend(aliases)

    out = {k: list(set(v)) for k, v in synonyms.items()}
    for std, aliases in newmap.items():
        out.setdefault(std, [])
        for a in aliases:
            if a not in out[std]:
                out[std].append(a)

    print("[OK] โหลด datadict:", len(newmap), "standards")
    return out


# -------------------------------------------------------------------
# อ่าน Excel แบบอัตโนมัติ (หาหัวตาราง)
# -------------------------------------------------------------------
def read_one_excel(path):
    """อ่าน Excel ด้วย pandas"""
    xl = pd.ExcelFile(path)
    sh = xl.sheet_names[0]
    for want in PREFERRED_SHEETS:
        for s in xl.sheet_names:
            if want in s:
                sh = s
                break
    raw = xl.parse(sh, header=None)
    h = detect_header_row(raw)
    df = raw.iloc[h + 1:].reset_index(drop=True)
    df.columns = make_unique([normalize_text(x) for x in raw.iloc[h].tolist()])
    return df


# -------------------------------------------------------------------
# ฟังก์ชันหลัก
# -------------------------------------------------------------------
def main(argv=None):
    cfg = yaml.safe_load(open(CFG_PATH, "r", encoding="utf-8"))
    synonyms = load_datadict_to_synonyms(cfg, COLUMN_SYNONYMS)

    input_file = cfg.get("input_file", os.path.join(DATA_DIR, "clean_data.xlsx"))
    output_file = cfg.get("output_file", os.path.join(OUT_DIR, "clean_with_formulas.xlsx"))
    zero_policy = cfg.get("zero_policy", "nan_to_zero")

    frames = []
    if os.path.exists(input_file):
        df = pd.read_excel(input_file)
        df = map_columns_flex(df, synonyms)
        frames.append(df)
    else:
        ignore = set(cfg.get('ignore_files', []))
    for fn in os.listdir(DATA_DIR):
        if fn in ignore:
            print(f'[INFO] ข้ามไฟล์ตาม ignore list: {fn}')
            continue
            if fn.endswith(".xlsx") and not fn.startswith("~$"):
                path = os.path.join(DATA_DIR, fn)
                try:
                    df = read_one_excel(path)
                    df = map_columns_flex(df, synonyms)
                    frames.append(df)
                except Exception as e:
                    print("[WARN] อ่านไม่ได้:", fn, e)

    if not frames:
        print("[ERROR] ไม่มีข้อมูล")
        return

    ALL = pd.concat(frames, ignore_index=True)

    # คำนวณคอลัมน์เพิ่ม
    for spec in cfg.get("computed_columns", []):
        name = spec.get("name")
        expr = spec.get("expr")
        rnd = spec.get("round", 2)
        try:
            val = eval_expr_df(ALL, expr)
            ALL[name] = pd.to_numeric(val, errors="coerce").round(rnd)
            print("[OK] computed:", name)
        except Exception as e:
            print("[WARN] ข้ามสูตร:", name, "->", e)

    # สร้างคอลัมน์ระดับบน
    if "แผนงาน" in ALL.columns:
        ALL["ระดับบน(ผลผลิตหรือโครงการ)"] = ALL["แผนงาน"].astype(str) + ":" + ALL.get("ผลผลิต/โครงการ", "")
    else:
        ALL["ระดับบน(ผลผลิตหรือโครงการ)"] = ""

    ALL.to_excel(output_file, index=False)
    print("[OK] wrote:", output_file, "rows:", len(ALL))


# -------------------------------------------------------------------
if __name__ == "__main__":
    main()
